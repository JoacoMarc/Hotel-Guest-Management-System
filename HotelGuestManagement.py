import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime
from colorama import init
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk
import os

# Initialize colorama
init()

class HotelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Luxor Hotel Management System")

        self.style = ttk.Style()
        self.style.configure("Red.TButton", background="red", foreground="white")
        
        self.id_list = []
        self.matrix = self.generate_matrix()
        
        if os.path.exists("guests.xlsx"):
            self.load_guest_file()

        self.create_widgets()

    def generate_matrix(self, rows=10, columns=6):
        return [[0 for _ in range(columns)] for _ in range(rows)]

    def validate_dates(self, start_date, end_date):
        start = datetime.strptime(str(start_date), '%d%m%Y')
        end = datetime.strptime(str(end_date), '%d%m%Y')
        return start < end

    def assign_room(self, id, first_name, last_name, birth_date, check_in, check_out, occupants):
        guest = {"id": id, "first_name": first_name, "last_name": last_name, 
                 "birth_date": birth_date, "check_in": check_in, "check_out": check_out, 
                 "occupants": occupants, "floor": None, "room": None}

        if occupants > 1:
            additional_occupants = []
            for i in range(occupants - 1):
                occupant_id = int(self.get_user_input(f"Enter ID for occupant {i+2}:"))
                occupant_first_name = self.get_user_input(f"Enter first name for occupant {i+2}:")
                occupant_last_name = self.get_user_input(f"Enter last name for occupant {i+2}:")
                occupant_birth_date = int(self.get_user_input(f"Enter birth date (ddmmyyyy) for occupant {i+2}:"))
                additional_occupants.append({"id": occupant_id, "first_name": occupant_first_name, 
                                             "last_name": occupant_last_name, "birth_date": occupant_birth_date})
            guest["additional_occupants"] = additional_occupants
        else:
            guest["additional_occupants"] = []

        while True:
            row = int(self.get_user_input("Enter floor (1-10):")) - 1
            col = int(self.get_user_input("Enter room number (1-6):")) - 1
            if self.matrix[row][col] == 0:
                guest["floor"] = row + 1
                guest["room"] = col + 1
                self.matrix[row][col] = guest
                break
            else:
                messagebox.showerror("Error", "Room already occupied, please choose another.")

        return self.matrix

    def get_user_input(self, prompt):
        return simpledialog.askstring("Input", prompt)

    def find_guest(self, last_name):
        for floor_index, floor in enumerate(self.matrix):
            for room_index, room in enumerate(floor):
                if room != 0 and room["last_name"].lower() == last_name.lower():
                    return (floor_index, room_index, room)
        return None

    def print_guest_info(self, guest):
        guest_info = (f"Guest Information:\n"
                      f"  ID: {guest['id']}\n"
                      f"  First Name: {guest['first_name']}\n"
                      f"  Last Name: {guest['last_name']}\n"
                      f"  Birth Date: {guest['birth_date']}\n"
                      f"  Check-in Date: {guest['check_in']}\n"
                      f"  Check-out Date: {guest['check_out']}\n"
                      f"  Number of Occupants: {guest['occupants']}\n"
                      f"  Floor: {guest['floor']}\n"
                      f"  Room: {guest['room']}")
        
        if guest["additional_occupants"]:
            guest_info += "\n  Additional Occupants:"
            for i, occupant in enumerate(guest["additional_occupants"], start=1):
                guest_info += (f"\n    Occupant {i}:\n"
                               f"      ID: {occupant['id']}\n"
                               f"      First Name: {occupant['first_name']}\n"
                               f"      Last Name: {occupant['last_name']}\n"
                               f"      Birth Date: {occupant['birth_date']}")
        
        messagebox.showinfo("Guest Info", guest_info)

    def most_occupied_floor(self):
        occupancy = [sum(1 for room in floor if room != 0) for floor in self.matrix]
        most_occupied = occupancy.index(max(occupancy))
        return most_occupied + 1

    def empty_rooms(self):
        return sum(1 for floor in self.matrix for room in floor if room == 0)

    def floor_with_most_occupants(self):
        occupants = [sum(room["occupants"] for room in floor if room != 0) for floor in self.matrix]
        return occupants.index(max(occupants)) + 1

    def next_departures(self, current_date):
        current = datetime.strptime(str(current_date), '%d%m%Y')
        departures = [(floor, room, datetime.strptime(str(room["check_out"]), '%d%m%Y')) 
                      for floor in self.matrix for room in floor if room != 0]

        if not departures:
            return []

        days_diff = [(floor, room, (departure - current).days) 
                     for floor, room, departure in departures]

        closest_days = min(days_diff, key=lambda x: x[2])[2]
        next_rooms = [room for floor, room, days in days_diff if days == closest_days]

        return next_rooms

    def create_guest_file(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Guests"

            headers = ["Floor", "Room", "ID", "First Name", "Last Name", "Birth Date", "Check-in Date", "Check-out Date", "Occupants"]

            ws.append(headers)

            for floor in self.matrix:
                for guest in floor:
                    if guest != 0:
                        guest_list = [guest["floor"], guest["room"], guest["id"], guest["first_name"], guest["last_name"], guest["birth_date"], 
                                      guest["check_in"], guest["check_out"], guest["occupants"]]
                        ws.append(guest_list)
                        for occupant in guest["additional_occupants"]:
                            occupant_list = [guest["floor"], guest["room"], occupant["id"], occupant["first_name"], occupant["last_name"], 
                                             occupant["birth_date"], "", "", ""]
                            ws.append(occupant_list)

            wb.save("guests.xlsx")
            messagebox.showinfo("Success", "Guest file created successfully.")

        except FileNotFoundError:
            messagebox.showerror("Error", "File not found.")
        except OSError as msg:
            messagebox.showerror("Error", f"Error: {msg}")

    def load_guest_file(self):
        try:
            wb = load_workbook("guests.xlsx")
            ws = wb.active
            self.matrix = self.generate_matrix()  # Reset matrix
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                floor, room, id, first_name, last_name, birth_date, check_in, check_out, occupants = row
                if id:
                    guest = {
                        "id": id,
                        "first_name": first_name,
                        "last_name": last_name,
                        "birth_date": birth_date,
                        "check_in": check_in,
                        "check_out": check_out,
                        "occupants": occupants,
                        "floor": floor,
                        "room": room,
                        "additional_occupants": []
                    }
                    self.matrix[floor - 1][room - 1] = guest
                    self.id_list.append(id)

            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                floor, room, id, first_name, last_name, birth_date, _, _, _ = row
                if id is None:
                    continue
                guest = self.matrix[floor - 1][room - 1]
                if guest != 0 and id != guest["id"]:
                    guest["additional_occupants"].append({
                        "id": id,
                        "first_name": first_name,
                        "last_name": last_name,
                        "birth_date": birth_date
                    })

        except FileNotFoundError:
            messagebox.showerror("Error", "Guest file not found.")
        except OSError as msg:
            messagebox.showerror("Error", f"Error: {msg}")

    def clear_guest_file(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Guests"

            headers = ["Floor", "Room", "ID", "First Name", "Last Name", "Birth Date", "Check-in Date", "Check-out Date", "Occupants"]
            ws.append(headers)

            wb.save("guests.xlsx")
            messagebox.showinfo("Success", "Guest file cleared successfully.")

        except FileNotFoundError:
            messagebox.showerror("Error", "File not found.")
        except OSError as msg:
            messagebox.showerror("Error", f"Error: {msg}")

    def enter_guest(self):
        try:
            id = int(self.get_user_input("Enter ID: "))
            assert len(str(id)) == 8, "ID must have 8 digits."
            assert id not in self.id_list, "ID already exists, please enter another."
            self.id_list.append(id)

            first_name = self.get_user_input("Enter first name:")
            last_name = self.get_user_input("Enter last name:")
            birth_date = int(self.get_user_input("Enter birth date (ddmmyyyy):"))
            check_in = int(self.get_user_input("Enter check-in date (ddmmyyyy):"))
            check_out = int(self.get_user_input("Enter check-out date (ddmmyyyy):"))
            assert self.validate_dates(check_in, check_out), "Check-in date must be before check-out date."
                
            occupants = int(self.get_user_input("Enter number of occupants:"))

            self.assign_room(id, first_name, last_name, birth_date, check_in, check_out, occupants)
            self.create_guest_file()  # Update the Excel file after adding a new guest

        except ValueError:
            messagebox.showerror("Error", "Please enter valid numbers for ID, birth date, check-in, and check-out dates.")
        except AssertionError as msg:
            messagebox.showerror("Error", str(msg))
            if id in self.id_list:
                self.id_list.remove(id)

    def find_guest_ui(self):
        last_name = self.get_user_input("Enter last name to search:")
        result = self.find_guest(last_name)
        if result:
            floor, room, guest = result
            self.print_guest_info(guest)
        else:
            messagebox.showerror("Error", "Guest not found.")

    def most_occupied_floor_ui(self):
        floor = self.most_occupied_floor()
        messagebox.showinfo("Most Occupied Floor", f"The most occupied floor is: {floor}")

    def empty_rooms_ui(self):
        empty_rooms_count = self.empty_rooms()
        messagebox.showinfo("Empty Rooms", f"The number of empty rooms is: {empty_rooms_count}")

    def floor_with_most_occupants_ui(self):
        floor = self.floor_with_most_occupants()
        messagebox.showinfo("Floor with Most Occupants", f"The floor with the most occupants is: {floor}")

    def next_departures_ui(self):
        current_date = int(self.get_user_input("Enter current date (ddmmyyyy):"))
        upcoming_rooms = self.next_departures(current_date)
        
        if not upcoming_rooms:
            messagebox.showinfo("Next Departures", "No upcoming departures found.")
            return
        
        departures_info = "Rooms with upcoming departures:\n"
        for room in upcoming_rooms:
            floor = room["floor"]
            room_number = room["room"]
            last_name = room["last_name"]
            departures_info += f"{last_name}, Floor: {floor}, Room: {room_number}\n"
        
        messagebox.showinfo("Next Departures", departures_info)

    def view_hotel(self):
        hotel_window = tk.Toplevel(self.root)
        hotel_window.title("Hotel View")

        canvas = tk.Canvas(hotel_window, width=800, height=600, bg="white")
        canvas.pack(expand=True, fill="both")

        room_width = 100
        room_height = 50
        padding = 10

        for i in range(10):
            floor_text = f"Floor {10 - i}"
            canvas.create_text(padding, (i + 1) * (room_height + padding) - room_height // 2, anchor="w", font=('Arial', 12, 'bold'), text=floor_text, fill="black")
            
            for j in range(6):
                room_status = self.matrix[9 - i][j]
                room_text = f"Room {j+1}\n"
                if room_status == 0:
                    color = "green"
                    room_text += "Empty"
                else:
                    color = "red"
                    room_text += f"Occupied\n({room_status['occupants']})"

                x0 = (j + 1) * (room_width + padding) + padding
                y0 = i * (room_height + padding) + padding
                x1 = x0 + room_width
                y1 = y0 + room_height

                canvas.create_rectangle(x0, y0, x1, y1, fill=color)
                canvas.create_text((x0 + x1) // 2, (y0 + y1) // 2, text=room_text, font=('Arial', 10, 'bold'), fill="black")

        close_button = ttk.Button(hotel_window, text="Close", command=hotel_window.destroy, style="Red.TButton")
        close_button.pack(pady=10)

        self.exit_button.pack(pady=10)
    
    def create_widgets(self):
        try:
            self.img = Image.open("/Users/joaquinmarcoff/Downloads/HotelLuxorBanner.jpg")
            self.img = self.img.resize((400, 100), Image.LANCZOS)  # Resize image
            self.photo = ImageTk.PhotoImage(self.img)
            self.image_label = tk.Label(self.root, image=self.photo)
            self.image_label.pack(pady=10)
        except FileNotFoundError:
            messagebox.showerror("Error", "Image file not found. Please check the path and try again.")

        self.add_guest_button = tk.Button(self.root, text="Enter New Guest", command=self.enter_guest)
        self.add_guest_button.pack(pady=10)

        self.find_guest_button = tk.Button(self.root, text="Find Guest by Last Name", command=self.find_guest_ui)
        self.find_guest_button.pack(pady=10)

        self.most_occupied_floor_button = tk.Button(self.root, text="Most Occupied Floor", command=self.most_occupied_floor_ui)
        self.most_occupied_floor_button.pack(pady=10)

        self.empty_rooms_button = tk.Button(self.root, text="Number of Empty Rooms", command=self.empty_rooms_ui)
        self.empty_rooms_button.pack(pady=10)

        self.floor_with_most_occupants_button = tk.Button(self.root, text="Floor with Most Occupants", command=self.floor_with_most_occupants_ui)
        self.floor_with_most_occupants_button.pack(pady=10)

        self.next_departures_button = tk.Button(self.root, text="Next Departures", command=self.next_departures_ui)
        self.next_departures_button.pack(pady=10)

        self.view_hotel_button = tk.Button(self.root, text="View Hotel", command=self.view_hotel)
        self.view_hotel_button.pack(pady=10)

        self.create_guest_file_button = tk.Button(self.root, text="Create or Update Guest Excel", command=self.create_guest_file)
        self.create_guest_file_button.pack(pady=10)

        self.clear_guest_file_button = tk.Button(self.root, text="Clear Guest Excel and Vacate Hotel", command=self.clear_guest_file)
        self.clear_guest_file_button.pack(pady=10)
        
        self.exit_button = ttk.Button(self.root, text="Exit", command=self.root.quit, style="Red.TButton")
        self.exit_button.pack(pady=10)


def main():
    root = tk.Tk()
    app = HotelApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()


